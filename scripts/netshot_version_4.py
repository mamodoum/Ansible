#!/usr/bin/python
# -*- coding:utf-8 -*-

### Pour lancer le script, utiliser la commande suivante:
##  python3.5 <fichier>.py  ##

#Début du programme
import time
start = time.time()

#Les lignes suivantes permettent de vérifier quelle version de python est utilisée.
#La version minimum nécessaire pour que le script fonctionne est la 3.5.
#Si la version de python utilisé n'est pas au moins la 3.5, un message explicite 
#est affiché à l'utilisateur et le script est stoppé.
import sys
if sys.version_info < (3,5):
    print("\nVeuillez relancer le programme et utiliser au moins la version <<3.5>> de python")
    print("Exit...\n")
    exit()

#Librairie Openpyxl permettant de traiter les donnees avec Excel
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv
import re

print('\n')
print (' ________________________________________________ ')
print ('|        ____________    ____________            |')
print ('|        |               |                       |')
print ('|        |               |                       |')
print ('|        |____________   |   _________           |')
print ('|                    |   |           |           |')
print ('|                    |   |           |           |')
print ('|                    |   |           |           |')
print ('|        ____________|   |___________|           |')
print ('|________________________________________________|') 
print('\n')

#Ce script permet de comparer les données contenues dans deux fichiers excel en utilisant la librairie openpyxl
#Il y a un fichier source et un fichier export de netshot
#Le but est d'extraire les hostnames qui correspondent dans les deux fichiers et les mettre dans un nouveau fichier
#excel
#On récupère également les hostnames qui sont dans le fichier source mais ne se trouvent pas dans netshot

print("Script is running.")
print('\n')

#Motif permettant de matcher les extensions excel
pattern = r"(.*.xls*)"

#Il faut ajouter l'extension en saisissant le nom du fichier!
print("-----------! AJOUTER L'EXTENSION (exemple test.xlsx ou test.xlsm) !-----------")
print('\n')

#Saisir le nom du fichier excel source et le nom de la feuille concernée
source_file_name = input("Saisissez le nom du fichier Excel source: ")

#Cette boucle while permet de tester si l'utilisateur à ajouter l'extension du fichier
#Si ce n'est pas fait, le message "Veuillez saisir le nom du fichier suivi de l'extension"
#sera répété tant qu'il ne rajoute pas l'extension
#Une fois le bon input saisi, il peut spécifier le nom de la feuille excel
# .lower() permet de mettre le nom du fichier source en minuscule pour la comparaison avec le pattern
#afin que au cas où l'extension est saisie en majuscule (exemple: monFichier.XSLX) par l'utilisateur
#il n'y ait pas d'erreur de matching
while re.match(pattern, source_file_name.lower()) is None:
	print('\n')
	print("Veuillez saisir le nom du fichier suivi de l'extension")
	source_file_name = input("Saisissez le nom du fichier Excel source: ")
	if re.match(pattern, source_file_name.lower()) : break

source_file_sheetname = input("Saisissez le nom de la feuille Excel concernee: ")

#source_file_name = "odin_MATCH_netshot-export_20180615-151513.xlsx"
#source_file_sheetname = "Sheet"

print('\n')
print("Script is running..")
print('\n')

#On charge le fichier source
try:
	wb1 = load_workbook(filename = source_file_name)
	ws1 = wb1[source_file_sheetname]
except:
	print ("\n")
	print("Impossible d'ouvrir le fichier << "+source_file_name+" >> ou d'acceder a la feuille << "+source_file_sheetname+" >>")
	print("Veuillez relancer le programme et saisir les bonnes informations.")
	print ("Exit..")
	print ("\n")
	exit()

print("Le fichier source utilise est: ", source_file_name)

#Pour toutes les lignes ou le x[1].value = hostname n'est pas nul, on récupère
#sous forme de dictionnaire  le matricule et le hostname/usual_name; ce dictionnaire est ensuite ajoutée au tableau 
#tab_source
#On vérifie si le matricule est nul, on remplace la valeur nulle par N/A
tab_source = []
for x in ws1.rows:
	if  x[1].value != None:
		if x[0].value == None:
			x[0].value = "N/A"
			tableau1 = {
			'matricule' : x[0].value,
			'usual_name' : x[1].value
			}
			tab_source.append(tableau1)
		else:
			tableau1 = {
			'matricule' : x[0].value,
			'usual_name' : x[1].value
			}
			tab_source.append(tableau1)

print('\n')
print("Script is running...")
print('\n')

##Saisir le nom du fichier excel netshot et le nom de la feuille concernée
#netshot_file_name = input("Saisissez le nom du fichier Excel netshot: ")

##Cette boucle while permet de tester si l'utilisateur à ajouter l'extension du fichier
##Si ce n'est pas fait, le message "Veuillez saisir le nom du fichier suivi de l'extension"
##sera répété tant qu'il ne rajoute pas l'extension
##Une fois le bon input saisi, il peut spécifier le nom de la feuille excel
#while re.match(pattern, netshot_file_name) is None:
#	print('\n')
#	print("Veuillez saisir le nom du fichier suivi de l'extension")
#	netshot_file_name = input("Saisissez le nom du fichier Excel netshot: ")
#	if re.match(pattern, source_file_name.lower()) : break
#
#netshot_file_sheetname = input("Saisissez le nom de la feuille Excel concernee: ")

#On charge le fichier netshot
netshot_file_name = "netshot-export_20180615-151513.xlsx"
netshot_file_sheetname = "Devices"

try:
	wb2 = load_workbook(filename = netshot_file_name)
	ws2 = wb2[netshot_file_sheetname]
except:
	print ("\n")
	print("Impossible d'ouvrir le fichier << "+netshot_file_name+" >> ou d'acceder a la feuille << "+netshot_file_sheetname+" >>")
	print("Veuillez rentrer les bonnes informations netshot dans le script et relancer le programme.")
	print ("Exit..")
	print ("\n")
	exit()

print("Le fichier netshot utilise est: ", netshot_file_name)
print('\n')

#Pour toutes les lignes ou le x[1].value = hostname n'est pas nul, on récupère
#sous forme de dictionnaire le hostname/usual_name, l'adresse IP et la version de l'OS (family);
#ce dictionnaire est ensuite ajoutée au tableau tab_netshot
tab_netshot = []
for x in ws2.rows:
	if x[1].value != None:
		tableau2 = {
		'name' : x[1].value,
		'management_ip' : x[2].value,
		'family' : x[5].value,
		'sofware_version' : x[8].value
		}
		tab_netshot.append(tableau2)


print("Script is running....")
print('\n')

#Ces deux lignes suivantes permettent de récupérer le nom des fichiers d'entrée sans les extensions (.xlsx ou .xlsm)
filename_source = source_file_name[:-5]
filename_netshot = netshot_file_name[:-5]

#Activation du workbook qui représente le fichier excel de sortie
wb_final = Workbook()

#On définit le nom du fichier de sortie
#exemple: fichiers d'entrée --> test1, netshot
#fichier de sortie --> test1_X_netshot
wb_final_filename = filename_source+"_X_"+filename_netshot+".xlsx"

#Création de deux feuilles excel
#Dans la feuille MATCH, il y aura les hostnames qui sont a la fois dans la source et netshot
#Dans la feuille NO_MATCH, il y aura les hostnames qui sont dans la source mais pas dans netshot
ws1_final = wb_final.create_sheet("MATCH")
ws2_final = wb_final.create_sheet("NO_MATCH")
ws3_final = wb_final.create_sheet("NXOS_7.1")
ws4_final = wb_final.create_sheet("NXOS_&_OTHERS")

#On enleve les entetes des colonnes pour ne conserver que les données
tab_source.remove(tab_source[0])
tab_netshot.remove(tab_netshot[0])

#Creation d'entetes pour les deux fichiers de sortie
enTete1 = ["Matricule","Hostname", "Management_ip", "Family", "Sofware_version", "Ansible_host"]
enTete2 = ["Matricule", "Hostname"]
ws1_final.append(enTete1)
ws2_final.append(enTete2)
ws3_final.append(enTete1)
ws4_final.append(enTete1)

#Les lignes suivantes permettent de mettre la couleur de fond des cellules de la première ligne
#(ligne des enetes) de chaque feuille, en jaune (FFEE08)
#range(len(entete)) --> permet d'itérer en fonction de la longueur de la ligne d'entetes.
#ord() reçoit les caractères Unicode et chr() les produit.
#On aura en premier A+0 = A qui represente la lettre de la colonne, puis A+1=B, 
#et ainsi de suite jusqu'a la fin de la boucle
#le "1" est pour la ligne numéro 1 (ligne des enetes)
#exemple : sheet['A1'].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = "solid") --> colore
#la premiere cellule (A) de la ligne 1 en jaune.
jj = 0
for x in range(len(enTete1)):
	column_letter = chr(ord('A')+jj)+"1"
	ws1_final[column_letter].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = "solid")
	ws3_final[column_letter].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = "solid")
	ws4_final[column_letter].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = "solid")
	jj = jj + 1
ii = 0
for x in range(len(enTete2)):
	column_letter = chr(ord('A')+ii)+"1"
	ws2_final[column_letter].fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = "solid")
	ii = ii + 1

#Pour toutes les valeurs de famille de nexus et la version 7.1(4)N1(1) qui correspondent, on sauvegarde dans
#la feuille NXOS; sinon dans la feuille NXOS_&_OTHERS
#Les versions de nexus qu'on recherche
nx1 = "Nexus 5672UP"
nx2 = "Nexus 6001"
nx3 = "Nexus 5624Q"
nx4 = "Nexus 56128UP"
nx5 = "Nexus 5548"
#Software version qu'on recherche
soft_vers = "7.1(4)N1(1)"

#Pour toutes les hostnames de netshot qui correspondent aux hostnames du fichier source, 
#On récupére le matricule, hostname, l'adresse IP, la version de l'OS dans un tableau
#Création d'une derniere colonne ansible_host dont les valeurs sont poussées dans ansible
#Ce tableau est ensuite ajouté dans la feuille excel MATCH
#Pour toutes les valeurs de famille de nexus et la version 7.1(4)N1(1) qui correspondent, on sauvegarde dans
#NXOS; sinon dans NXOS_&_OTHERS
for x in tab_netshot:
	ansible_host = x['name']+" ansible_host="+x['management_ip']
	for y in tab_source:
		if str(x['name']) == str(y['usual_name']):
			result1 = [y['matricule'], x['name'], x['management_ip'], x['family'], x['sofware_version'], ansible_host]
			ws1_final.append(result1)
			if (x['family']==nx1) or (x['family']==nx2) or (x['family']==nx3) or (x['family']==nx4) or (x['family']==nx5 and x['sofware_version']==soft_vers):
				result3 = [y['matricule'], x['name'], x['management_ip'], x['family'], x['sofware_version'], ansible_host]
				ws3_final.append(result3)
			else:
				result4 = [y['matricule'], x['name'], x['management_ip'], x['family'], x['sofware_version'], ansible_host]
				ws4_final.append(result4)


#Pour toutes les hostnames du ficher source qui ne se trouvent pas dans netshot,
#on récupère les hostnames; le booléen "found" permet de tagguer les hostnames;
#les valeurs qui correspondent sont a true; ce qui permet de récupérer les valeurs 
#qui ne correspondent pas qui sont a false
#Ces valeurs sont ensuite ajoutées a la feuille NO_MATCH
for x in tab_source:
	found = False
	for y in tab_netshot:
		if x['usual_name'] == y['name']:
			found = True
	if found == False:
		result2 = [x['matricule'], x['usual_name']]
		ws2_final.append(result2)
	

#Enregistrement du fichier de sortie.
#Suppression de la feuille vide qui se crée par défaut lorsqu'on crée le workbook de sortie
wb_final.remove(wb_final["Sheet"])
wb_final.save(wb_final_filename)

print("Script is running.....")
print('\n')

print("Le fichier de sortie est: ", wb_final_filename)
print('\n')

print("\n######################################")
print("Script ..........Completed")
print("######################################")
print ('\nIt tooks', int((time.time()-start)/60), 'minutes',  float((time.time()-start)%60), 'seconds.\n')